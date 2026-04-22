# services/excel_service.py

import os
import csv
import traceback
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from django.conf import settings

class ExcelService:
    def __init__(self):
        self.xlsx_file = os.path.join(settings.BASE_DIR, 'leads.xlsx')
        self.csv_dir = os.path.join(settings.BASE_DIR, 'leads_csv')
        os.makedirs(self.csv_dir, exist_ok=True)

        self.sheets = ['Обратный звонок', 'Консультации', 'Контакты', 'Партнеры']
        self.headers = ['№', 'Имя клиента', 'Почта', 'Номер телефона', 'Описание', 'Дата']

        # Создаём CSV-файлы, если их нет
        self._init_csv()
        # Проверяем XLSX и создаём, если нужно
        self._ensure_xlsx()

    def _ensure_xlsx(self):
        """Создать XLSX файл с листами, если его нет"""
        if os.path.exists(self.xlsx_file):
            return
        try:
            wb = Workbook()
            wb.remove(wb.active)
            for sheet_name in self.sheets:
                ws = wb.create_sheet(sheet_name)
                ws.append(self.headers)
                ws.column_dimensions['A'].width = 6
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 50
                ws.column_dimensions['F'].width = 20
                # для будущих строк перенос текста будем включать при добавлении
            wb.save(self.xlsx_file)
            print(f"✅ XLSX файл создан: {self.xlsx_file}")
        except Exception as e:
            print(f"❌ Ошибка при создании XLSX: {e}")
            traceback.print_exc()

    def _init_csv(self):
        for sheet_name in self.sheets:
            file_path = os.path.join(self.csv_dir, f"{sheet_name}.csv")
            if not os.path.exists(file_path):
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f, quoting=csv.QUOTE_ALL)
                    writer.writerow(self.headers)

    def _get_next_row_number(self, ws):
        return ws.max_row + 1

    def _add_to_xlsx(self, sheet_name, row_data):
        """Добавить строку в XLSX (с обязательной проверкой и открытием)"""
        try:
            # Убеждаемся, что файл существует
            self._ensure_xlsx()

            # Открываем книгу (возможно, заблокирована, но попробуем)
            wb = load_workbook(self.xlsx_file)

            # Проверяем наличие листа
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(self.headers)
            else:
                ws = wb[sheet_name]

            # Получаем номер следующей строки
            row_num = self._get_next_row_number(ws)

            # Строка с номером
            row_data_with_num = [row_num] + row_data[1:]
            ws.append(row_data_with_num)

            # Включаем перенос текста для колонки E (Описание)
            for cell in ws[row_num]:
                if cell.column == 5:
                    cell.alignment = Alignment(wrap_text=True)

            # Сохраняем
            wb.save(self.xlsx_file)
            print(f"✅ XLSX запись добавлена в лист '{sheet_name}', строка {row_num}")
            return True
        except Exception as e:
            print(f"❌ Ошибка записи в XLSX: {e}")
            traceback.print_exc()
            return False

    def _add_to_csv(self, sheet_name, row_data):
        file_path = os.path.join(self.csv_dir, f"{sheet_name}.csv")
        try:
            with open(file_path, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, quoting=csv.QUOTE_ALL)
                writer.writerow(row_data)
            return True
        except Exception as e:
            print(f"❌ Ошибка записи в CSV: {e}")
            traceback.print_exc()
            return False

    def _add_record(self, sheet_name, record_data):
        xlsx_ok = self._add_to_xlsx(sheet_name, record_data)
        csv_ok = self._add_to_csv(sheet_name, record_data)
        return xlsx_ok or csv_ok

    # ----- Публичные методы -----
    def add_callback(self, name, phone):
        data = [0, name, '', phone, 'Заявка на обратный звонок', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ok = self._add_record('Обратный звонок', data)
        if ok:
            print(f"✅ Обратный звонок | {name} | {phone}")
        return ok

    def add_consultation(self, name, email, phone, message, bot_name=''):
        description = f"Консультация по боту: {bot_name}\nСообщение: {message}" if bot_name else f"Консультация\nСообщение: {message}"
        data = [0, name, email, phone, description, datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ok = self._add_record('Консультации', data)
        if ok:
            print(f"✅ Консультация | {name} | {email} | {phone}")
        return ok

    def add_contact(self, name, email, phone, message):
        data = [0, name, email, phone or '', message, datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ok = self._add_record('Контакты', data)
        if ok:
            print(f"✅ Контакт | {name} | {email} | {message[:50]}...")
        return ok

    def add_partner(self, name, email, company):
        description = f"Компания: {company}\nЗаявка на партнерство"
        data = [0, name, email, '', description, datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ok = self._add_record('Партнеры', data)
        if ok:
            print(f"✅ Партнер | {name} | {email} | {company}")
        return ok

excel_service = ExcelService()