import os
import time
import tempfile
import shutil
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)


class BaseExcelWriter:
    MAX_RETRIES = 3
    RETRY_DELAY = 0.5

    def __init__(self, base_dir, prefix, headers, backup_dir):
        self.base_dir = base_dir
        self.prefix = prefix
        self.headers = headers
        self.backup_dir = backup_dir

        os.makedirs(self.base_dir, exist_ok=True)
        os.makedirs(self.backup_dir, exist_ok=True)
        
        self.current_file_path = os.path.join(self.base_dir, f"{self.prefix}.xlsx")

    def _get_archive_path(self):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return os.path.join(self.backup_dir, f"backup_{timestamp}.xlsx")

    def _ensure_dir_exists(self, file_path):
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

    def _get_row_count(self, file_path, sheet_name):
        try:
            if not os.path.exists(file_path):
                return 0
            wb = load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                return 0
            ws = wb[sheet_name]
            return max(0, ws.max_row - 1)
        except Exception as e:
            logger.error(f"Ошибка при подсчёте строк: {e}")
            return 0

    def _apply_default_font(self, ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='Times New Roman', size=14, color="000000")

    def _setup_data_validation(self, ws):
        """Настраивает выпадающий список для колонки Статус заявки (колонка G, индекс 7)"""
        status_col = 7
        max_row = ws.max_row if ws.max_row > 1 else 1000
        status_range = f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{max_row}"
        
        # Удаляем старые валидации
        for dv in list(ws.data_validations.dataValidation):
            ws.data_validations.dataValidation.remove(dv)
        
        # Создаём новую валидацию
        dv = DataValidation(
            type="list",
            formula1='"Не обработано,Обработано"',
            allow_blank=False,
            showDropDown=True
        )
        dv.error = 'Выберите значение из списка'
        dv.errorTitle = 'Недопустимое значение'
        dv.prompt = 'Выберите статус заявки'
        dv.promptTitle = 'Выпадающий список'
        
        ws.add_data_validation(dv)
        dv.add(status_range)
        
        # Применяем стили ко всем строкам
        for row in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row, column=status_col)
            if status_cell.value == "Обработано":
                status_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            else:
                status_cell.value = "Не обработано"
                status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            status_cell.font = Font(name='Times New Roman', size=14, color="000000")
        
        logger.info(f"✅ Настроен выпадающий список для колонки {get_column_letter(status_col)}")

    def auto_fit_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell.row == 1:
                            cell_length = min(cell_length + 2, 60)
                        max_length = max(max_length, cell_length)
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[col_letter].width = adjusted_width

    def create_file(self, sheet_name, file_path=None):
        """Создаёт новый Excel файл с выпадающим списком для статуса"""
        if file_path is None:
            file_path = self.current_file_path
        
        self._ensure_dir_exists(file_path)
        
        # Если файл уже существует, просто настраиваем валидацию
        if os.path.exists(file_path):
            return self._setup_existing_file(file_path, sheet_name)
        
        for attempt in range(self.MAX_RETRIES):
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                ws.append(self.headers)
                
                self._apply_default_font(ws)
                for cell in ws[1]:
                    cell.font = Font(name='Times New Roman', size=14, bold=True)
                
                # Настраиваем выпадающий список
                self._setup_data_validation(ws)
                
                self.auto_fit_columns(ws)
                wb.save(file_path)
                logger.info(f"✅ Создан файл: {file_path}")
                return True
                
            except PermissionError:
                logger.warning(f"⚠️ Файл заблокирован, попытка {attempt + 1}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
            except Exception as e:
                logger.error(f"❌ Ошибка создания: {e}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
        return False

    def _setup_existing_file(self, file_path, sheet_name):
        """Настраивает выпадающий список в существующем файле"""
        for attempt in range(self.MAX_RETRIES):
            try:
                with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False, dir=os.path.dirname(file_path)) as tmp_file:
                    tmp_path = tmp_file.name
                
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                    ws.append(self.headers)
                    for cell in ws[1]:
                        cell.font = Font(name='Times New Roman', size=14, bold=True)
                else:
                    ws = wb[sheet_name]
                
                # Настраиваем выпадающий список
                self._setup_data_validation(ws)
                
                self.auto_fit_columns(ws)
                wb.save(tmp_path)
                shutil.move(tmp_path, file_path)
                logger.info(f"✅ Обновлён файл: {file_path}")
                return True
                
            except PermissionError:
                logger.warning(f"⚠️ Файл заблокирован, попытка {attempt + 1}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
            except Exception as e:
                logger.error(f"❌ Ошибка: {e}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
        return False

    def add_row(self, sheet_name, row_data):
        """Добавляет строку в Excel файл"""
        target_file = self.current_file_path
        self._ensure_dir_exists(target_file)
        
        # Проверяем, существует ли файл, если нет — создаём
        if not os.path.exists(target_file):
            self.create_file(sheet_name, target_file)

        for attempt in range(self.MAX_RETRIES):
            try:
                with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False, dir=os.path.dirname(target_file)) as tmp_file:
                    tmp_path = tmp_file.name

                wb = load_workbook(target_file)
                
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                    ws.append(self.headers)
                    self._apply_default_font(ws)
                    for cell in ws[1]:
                        cell.font = Font(name='Times New Roman', size=14, bold=True)
                else:
                    ws = wb[sheet_name]

                # Добавляем строку с автоматическим номером
                row_num = ws.max_row + 1
                row_with_num = [row_num - 1] + list(row_data)
                ws.append(row_with_num)
                
                # Применяем шрифт к новой строке
                for cell in ws[row_num]:
                    cell.font = Font(name='Times New Roman', size=14, color="000000")
                
                # Применяем статус "Не обработано" с красным фоном
                status_col = 7
                if len(row_with_num) >= status_col:
                    status_cell = ws.cell(row=row_num, column=status_col)
                    status_cell.value = "Не обработано"
                    status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    status_cell.font = Font(name='Times New Roman', size=14, color="000000")
                
                # Обновляем выпадающий список для всего столбца
                self._setup_data_validation(ws)
                
                self.auto_fit_columns(ws)
                
                wb.save(tmp_path)
                wb.close()

                # Создаём резервную копию
                if os.path.exists(target_file):
                    backup_path = self._get_archive_path()
                    shutil.copy2(target_file, backup_path)

                shutil.move(tmp_path, target_file)
                logger.info(f"✅ Строка добавлена в '{sheet_name}'")
                return True

            except PermissionError:
                logger.warning(f"⚠️ Файл заблокирован, попытка {attempt + 1}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
            except Exception as e:
                logger.error(f"❌ Ошибка: {e}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
        return False

    def read_all_rows_with_styles(self, sheet_name):
        """Читает все строки из Excel файла"""
        file_path = self.current_file_path
        
        for attempt in range(self.MAX_RETRIES):
            try:
                if not os.path.exists(file_path):
                    return []
                
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    return []
                
                ws = wb[sheet_name]
                
                rows = []
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    row_data = list(row)
                    status = "Не обработано"
                    if len(row_data) >= 7:
                        cell_value = row_data[6]  # 7-я колонка
                        if cell_value == "Обработано":
                            status = "Обработано"
                    rows.append({
                        'row_num': row_idx,
                        'data': row_data,
                        'status': status
                    })
                return rows
                
            except PermissionError:
                logger.warning(f"⚠️ Файл заблокирован при чтении, попытка {attempt + 1}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
            except Exception as e:
                logger.error(f"❌ Ошибка чтения: {e}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
        return []

    def update_cell_style(self, sheet_name, row_num, col_num, value, fill_color=None):
        """Обновляет ячейку статуса и её цвет"""
        target_file = self.current_file_path
        self._ensure_dir_exists(target_file)

        for attempt in range(self.MAX_RETRIES):
            try:
                if not os.path.exists(target_file):
                    return False
                
                with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False, dir=os.path.dirname(target_file)) as tmp_file:
                    tmp_path = tmp_file.name

                wb = load_workbook(target_file)
                if sheet_name not in wb.sheetnames:
                    return False
                
                ws = wb[sheet_name]
                
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                
                # Устанавливаем цвет фона
                if value == "Обработано":
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                
                cell.font = Font(name='Times New Roman', size=14, color="000000")
                
                wb.save(tmp_path)
                wb.close()
                
                shutil.move(tmp_path, target_file)
                logger.info(f"✅ Обновлена ячейка: строка {row_num}, статус = {value}")
                return True
                
            except PermissionError:
                logger.warning(f"⚠️ Файл заблокирован, попытка {attempt + 1}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
            except Exception as e:
                logger.error(f"❌ Ошибка обновления: {e}")
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(self.RETRY_DELAY)
        return False

    def get_cell_value(self, sheet_name, row_num, column_name):
        """Получает значение ячейки"""
        target_file = self.current_file_path
        
        try:
            if not os.path.exists(target_file):
                return ''
            
            wb = load_workbook(target_file)
            if sheet_name not in wb.sheetnames:
                return ''
            
            ws = wb[sheet_name]
            
            headers_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            col_idx = None
            for idx, header in enumerate(headers_row):
                if header == column_name:
                    col_idx = idx + 1
                    break
            
            if col_idx is None:
                return ''
            
            cell = ws.cell(row=row_num, column=col_idx)
            return cell.value if cell.value else ''
            
        except Exception as e:
            logger.error(f"Ошибка получения значения: {e}")
            return ''
