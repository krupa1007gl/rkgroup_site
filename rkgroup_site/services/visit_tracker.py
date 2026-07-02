import os
import logging
from datetime import datetime, timedelta
from collections import Counter
from django.conf import settings
from .excel_base import BaseExcelWriter

logger = logging.getLogger(__name__)


class VisitTracker:
    def __init__(self):
        self.headers = ['№', 'IP', 'Время', 'URL', 'Referer', 'User-Agent', 'Статус']
        self.visits_dir = settings.VISITS_DIR
        self.backups_dir = settings.BACKUPS_DIR
        os.makedirs(self.visits_dir, exist_ok=True)

        self.writer = BaseExcelWriter(
            base_dir=self.visits_dir,
            prefix='visits',
            headers=self.headers,
            backup_dir=self.backups_dir
        )
        self._ensure_current_file()
        logger.info("VisitTracker инициализирован")

    def _ensure_current_file(self):
        self.writer.create_file('Посещения')

    def add_visit(self, ip, url, referer, user_agent, status):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [ip, timestamp, url, referer or '-', user_agent or '-', status]
        return self.writer.add_row('Посещения', row_data)

    def get_stats_for_period(self, start_date, end_date):
        all_rows = []
        current = start_date.replace(day=1)
        while current <= end_date:
            year_month = current.strftime('%Y-%m')
            rows = self.writer.read_all_rows_with_styles('Посещения', current)
            for row in rows:
                all_rows.append(row['data'])
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        if not all_rows:
            return self._empty_stats()

        filtered = []
        for row in all_rows:
            if len(row) < 7:
                continue
            try:
                row_date = datetime.strptime(row[2], '%Y-%m-%d %H:%M:%S')
                if start_date <= row_date <= end_date:
                    filtered.append(row)
            except:
                continue

        if not filtered:
            return self._empty_stats()

        total_visits = len(filtered)
        unique_ips = len(set(row[1] for row in filtered))
        pages = Counter(row[3] for row in filtered if row[3])
        referers = Counter(row[4] for row in filtered if row[4] and row[4] != '-')

        return {
            'total_visits': total_visits,
            'unique_ips': unique_ips,
            'top_pages': [{'url': url, 'count': count} for url, count in pages.most_common(10)],
            'top_referers': [{'referer': ref, 'count': count} for ref, count in referers.most_common(10)],
        }

    def _empty_stats(self):
        return {'total_visits': 0, 'unique_ips': 0, 'top_pages': [], 'top_referers': []}

    def export_stats_to_excel(self, start_date, end_date, stats):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Общая статистика"
        ws_summary['A1'] = "Период"
        ws_summary['B1'] = f"{start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"
        ws_summary['A3'] = "Показатель"
        ws_summary['B3'] = "Значение"
        ws_summary['A4'] = "Всего визитов"
        ws_summary['B4'] = stats['total_visits']
        ws_summary['A5'] = "Уникальных IP"
        ws_summary['B5'] = stats['unique_ips']

        for cell in ['A1', 'A3', 'B3']:
            ws_summary[cell].font = Font(bold=True)

        ws_pages = wb.create_sheet("Топ-10 страниц")
        ws_pages['A1'] = "URL"
        ws_pages['B1'] = "Количество"
        for cell in ['A1', 'B1']:
            ws_pages[cell].font = Font(bold=True)
        for idx, page in enumerate(stats['top_pages'], start=2):
            ws_pages[f'A{idx}'] = page['url']
            ws_pages[f'B{idx}'] = page['count']

        ws_refs = wb.create_sheet("Топ-10 источников")
        ws_refs['A1'] = "Referer"
        ws_refs['B1'] = "Количество"
        for cell in ['A1', 'B1']:
            ws_refs[cell].font = Font(bold=True)
        for idx, ref in enumerate(stats['top_referers'], start=2):
            ws_refs[f'A{idx}'] = ref['referer']
            ws_refs[f'B{idx}'] = ref['count']

        for ws in [ws_summary, ws_pages, ws_refs]:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)

        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()
        return tmp.name


visit_tracker = VisitTracker()
