import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

FORMULA_TRIGGER_CHARS = ('=', '+', '-', '@')


def escape_formula(value):
    """
    Защита от CSV/Excel-инъекции формул: если строковое значение начинается
    с одного из спецсимволов, которые Excel/LibreOffice интерпретируют как
    начало формулы, добавляем ведущий апостроф — Excel покажет значение как
    текст, формула не выполнится.
    """
    if isinstance(value, str) and value.startswith(FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)


def build_workbook(sheets):
    """
    Собирает xlsx-файл в памяти из списка листов.
    sheets: список (sheet_name, headers, rows), rows — список списков значений.
    Все значения экранируются от формул перед записью.
    Возвращает BytesIO, готовый для HttpResponse.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, headers, rows in sheets:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([escape_formula(v) for v in row])
        _autofit(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
