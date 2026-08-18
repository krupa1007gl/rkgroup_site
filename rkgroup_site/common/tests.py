from django.core.cache import cache
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from common.mixins import check_rate_limit_key
from services.xlsx_export import build_workbook, escape_formula


class EscapeFormulaTests(TestCase):
    def test_escapes_leading_equals(self):
        self.assertEqual(escape_formula('=SUM(A1:A9)'), "'=SUM(A1:A9)")

    def test_escapes_leading_plus_minus_at(self):
        self.assertEqual(escape_formula('+1234'), "'+1234")
        self.assertEqual(escape_formula('-1234'), "'-1234")
        self.assertEqual(escape_formula('@cmd'), "'@cmd")

    def test_leaves_normal_strings_untouched(self):
        self.assertEqual(escape_formula('Иван Иванов'), 'Иван Иванов')

    def test_leaves_non_strings_untouched(self):
        self.assertEqual(escape_formula(42), 42)
        self.assertEqual(escape_formula(None), None)


class BuildWorkbookTests(TestCase):
    def test_formula_injection_is_escaped_in_output_file(self):
        buffer = build_workbook([
            ('Заявки', ['Имя', 'Описание'], [['Иван', '=SUM(A1:A9)'], ['Пётр', '+79161234567']]),
        ])
        wb = load_workbook(buffer)
        ws = wb['Заявки']
        self.assertEqual(ws.cell(row=2, column=2).value, "'=SUM(A1:A9)")
        self.assertEqual(ws.cell(row=3, column=2).value, "'+79161234567")


class RateLimitHelperTests(TestCase):
    def setUp(self):
        cache.clear()

    @override_settings(RATELIMIT_ENABLED=True)
    def test_allows_up_to_limit_then_blocks(self):
        key = 'test-rl-key'
        for _ in range(3):
            self.assertTrue(check_rate_limit_key(key, limit=3))
        self.assertFalse(check_rate_limit_key(key, limit=3))

    @override_settings(RATELIMIT_ENABLED=False)
    def test_disabled_never_blocks(self):
        key = 'test-rl-key-disabled'
        for _ in range(10):
            self.assertTrue(check_rate_limit_key(key, limit=1))
